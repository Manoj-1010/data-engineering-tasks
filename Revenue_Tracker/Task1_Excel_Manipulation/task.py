import pandas as pd
from openpyxl import load_workbook
 
def to_number(x):
    try:
        if isinstance(x, str):
            x = x.replace("£", "").replace(",", "").strip()
        return float(x)
    except (ValueError, TypeError):
        return 0.0
 
month_to_column = {
    "April": "D", "May": "E", "June": "F", "July": "G",
    "August": "H", "September": "I", "October": "J",
    "November": "K", "December": "L", "January": "M",
    "February": "N", "March": "O"
}
 
INPUT_FILE = "Delta3_Apr.xlsx"
TEMPLATE_FILE = "Delta3_Output.xlsx"
 
ROW_REVENUE = 7
ROW_REVENUE_PCT = 8
ROW_WORKFORCE = 10
ROW_WORKFORCE_PCT = 11
ROW_ALLOC = 13
ROW_ALLOC_PCT = 14
 
if __name__ == "__main__":
 
    xls = pd.ExcelFile(INPUT_FILE)
    wb = load_workbook(TEMPLATE_FILE)
    ws = wb.active

    for sheet in xls.sheet_names:
 
        if sheet not in month_to_column:
            continue
 
        column = month_to_column[sheet]
 
        df = xls.parse(sheet, header=None).dropna(how='all')
        data_map = df.set_index(0)[1].to_dict()
 
        revenue = to_number(data_map.get("Revenue", 0))
 
        salary_core = to_number(data_map.get("Salary - Core employees", 0))
        salary_mgr = to_number(data_map.get("Salary - TL / Managers", 0))
        salary_cons = to_number(data_map.get("Salary - Consultants", 0))
        incentives = to_number(data_map.get("Performance payments - Incentive & Others", 0))
 
        total_workforce_cost = salary_core + salary_mgr + salary_cons + incentives
 
        ws[f"{column}{ROW_REVENUE}"].value = revenue
        ws[f"{column}{ROW_REVENUE_PCT}"].value = data_map.get("Revenue %", 0)
 
        ws[f"{column}{ROW_WORKFORCE}"].value = total_workforce_cost
        ws[f"{column}{ROW_WORKFORCE_PCT}"].value = (
            total_workforce_cost / revenue if revenue else 0
        )
 
        ws[f"{column}{ROW_ALLOC}"].value = data_map.get("Total salary allocation for project", 0)
        ws[f"{column}{ROW_ALLOC_PCT}"].value = data_map.get("Total salary allocation %", 0)

        ws[f"P{ROW_REVENUE}"].value  =  to_number(ws[f"P{ROW_REVENUE}"].value) + revenue
        ws[f"P{ROW_WORKFORCE}"].value  =  to_number(ws[f"P{ROW_WORKFORCE}"].value) + total_workforce_cost
        ws[f"P{ROW_ALLOC}"].value  +=  to_number(ws[f"{column}{ROW_ALLOC}"].value)
 
    wb.save("Output.xlsx")
